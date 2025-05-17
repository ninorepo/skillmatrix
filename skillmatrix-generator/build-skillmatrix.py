import openpyxl
import sys



# Load workbooks
source_file_path = sys.argv[1]
output_file = sys.argv[2]

source_file = openpyxl.load_workbook(source_file_path , data_only=True)
template_wb = openpyxl.load_workbook("template_skillmatrix.xlsx")
source_ws = source_file.active
template_ws = template_wb.active

# 1. Date (B1) → A2
template_ws["A2"].value = source_ws["B1"].value

# 2. Line (B2) → A1 as "SKILL MATRIX <line>"
template_ws["A1"].value = f"SKILL MATRIX {source_ws['B2'].value}"

# 3. Coordinate setup
row = 13
current_name_row = 4
current_operation_col = 5
name_to_row = {}
operation_to_col = {}

# 4. Transfer data
while True:
    id_val = source_ws[f"A{row}"].value
    name_val = source_ws[f"B{row}"].value
    operation_val = source_ws[f"C{row}"].value
    machine_val = source_ws[f"H{row}"].value
    performance_val = source_ws[f"U{row}"].value

    if id_val is None:
        break

    if name_val and "—" in str(name_val):
        row += 1
        continue

    # Register name
    if name_val not in name_to_row:
        template_ws.cell(row=current_name_row, column=2).value = id_val
        template_ws.cell(row=current_name_row, column=3).value = name_val
        name_to_row[name_val] = current_name_row
        current_name_row += 1

    # Register operation
    if operation_val not in operation_to_col:
        template_ws.cell(row=3, column=current_operation_col).value = operation_val  # Skip row 2
        operation_to_col[operation_val] = current_operation_col
        current_operation_col += 1

    # Place performance
    r = name_to_row[name_val]
    c = operation_to_col[operation_val]
    template_ws.cell(row=r, column=c).value = performance_val

    row += 1

# 5. Delete extra rows that don't contain ID and NAME, skipping the TOTAL row
for row in range(template_ws.max_row, 3, -1):
    id_cell = template_ws.cell(row=row, column=2).value
    name_cell = template_ws.cell(row=row, column=3).value
    total_cell = template_ws.cell(row=row, column=1).value

    if total_cell and isinstance(total_cell, str) and "TOTAL" in total_cell.upper():
        continue

    if not id_cell or not name_cell:
        template_ws.delete_rows(row)

# 6. Delete extra columns that don't contain operations
for col in range(template_ws.max_column, 5, -1):
    has_operation = False
    for row in range(3, 4):
        if template_ws.cell(row=row, column=col).value:
            has_operation = True
            break
    if not has_operation:
        template_ws.delete_cols(col)

# 7. Copy machine codes from H13 downward → E1 to the right, skipping "—"
source_row = 13
target_col = 5  # Column E
while True:
    val = source_ws[f"H{source_row}"].value
    if val is None:
        break
    if str(val).strip() != "—":
        template_ws.cell(row=1, column=target_col).value = val
        target_col += 1
    source_row += 1

# 8. Save the file
template_wb.save(output_file)

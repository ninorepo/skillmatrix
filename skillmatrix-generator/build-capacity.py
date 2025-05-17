import openpyxl
import copy
from openpyxl.styles import Alignment, Border, Side
import sys

# Load the source and template files
source_file_path =  sys.argv[1]
template_file_path = 'template_capacity.xlsx'

source_file = openpyxl.load_workbook(source_file_path)
template_workbook = openpyxl.load_workbook(template_file_path)
output_file = sys.argv[2]


# Access the first sheet of each workbook
source_sheet = source_file.active
template_sheet = template_workbook.active

# Create a new workbook and copy the template structure
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = template_sheet.title

# Copy the template structure including formatting
for row in template_sheet.iter_rows():
    for cell in row:
        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
        new_cell.font = copy.copy(cell.font)
        new_cell.fill = copy.copy(cell.fill)
        new_cell.border = copy.copy(cell.border)
        new_cell.number_format = copy.copy(cell.number_format)

        # Set wrap_text
        existing_alignment = cell.alignment or Alignment()
        new_cell.alignment = Alignment(
            horizontal=existing_alignment.horizontal,
            vertical=existing_alignment.vertical,
            wrap_text=True
        )

# Copy merged cells
for merged_range in template_sheet.merged_cells.ranges:
    new_sheet.merge_cells(str(merged_range))

# Copy column widths
for col_letter, dim in template_sheet.column_dimensions.items():
    new_sheet.column_dimensions[col_letter].width = dim.width

# Copy row heights
for row_idx, dim in template_sheet.row_dimensions.items():
    new_sheet.row_dimensions[row_idx].height = dim.height

# Copy the first 9 rows of metadata
new_sheet['B1'] = source_sheet['B1'].value
new_sheet['D1'] = source_sheet['B2'].value
new_sheet['G1'] = source_sheet['B9'].value
new_sheet['I1'] = source_sheet['B3'].value
new_sheet['M1'] = source_sheet['B4'].value
new_sheet['Q1'] = source_sheet['B5'].value
new_sheet['T1'] = source_sheet['B6'].value
new_sheet['D75'] = source_sheet['B7'].value
new_sheet['D76'] = source_sheet['B8'].value

# Copy data rows from source A13+ to destination A4+
for row in range(13, source_sheet.max_row + 1):
    for col in range(1, source_sheet.max_column + 1):
        value = source_sheet.cell(row=row, column=col).value
        new_cell = new_sheet.cell(row=row - 9, column=col, value=value)

        # Set wrap_text
        existing_alignment = new_cell.alignment or Alignment()
        new_cell.alignment = Alignment(
            horizontal=existing_alignment.horizontal,
            vertical=existing_alignment.vertical,
            wrap_text=True
        )

# Delete empty rows
rows_to_delete = []
for row in range(1, new_sheet.max_row + 1):
    if all(new_sheet.cell(row=row, column=col).value in (None, "") for col in range(1, new_sheet.max_column + 1)):
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    new_sheet.delete_rows(row)

# Apply outer box border (only on the outside of the row) for the entire row (columns 1 to 24) if column B contains "—"
outer_border = Border(
    left=Side(border_style='thick'),
    right=Side(border_style='thick'),
    top=Side(border_style='thick'),
    bottom=Side(border_style='thick')
)

# Starting from row 4 (after metadata/header)
for row in range(4, new_sheet.max_row + 1):
    cell_value = new_sheet.cell(row=row, column=2).value  # Column B
    if isinstance(cell_value, str) and "—" in cell_value:
        # Apply outer thick box border for columns 1 to 24 in this row
        # Apply left border of the first column and right border of the last column
        new_sheet.cell(row=row, column=1).border = Border(left=Side(border_style='thick'))  # Left border of first column
        new_sheet.cell(row=row, column=24).border = Border(right=Side(border_style='thick'))  # Right border of 24th column
        
        # Apply top and bottom borders across columns 1 to 24 (for the entire row)
        for col in range(1, 25):  # Columns 1 to 24
            new_sheet.cell(row=row, column=col).border = Border(top=Side(border_style='thick'), bottom=Side(border_style='thick'))

# Save the new workbook
new_workbook.save(output_file)

print(f"New file created and saved at: {output_file}")

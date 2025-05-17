from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys

def parse_number_if_numeric(value):
    if isinstance(value, str):
        value = value.strip()
        # Check if it's a valid int or float (but not alphanumeric like "123A")
        try:
            return int(value) if value.isdigit() else float(value)
        except ValueError:
            return value  # Not a number, keep as string
    return value  # Already numeric or other type



# File paths
source_file = sys.argv[1]
template_file = "template-individual.xlsx"
output_file =  sys.argv[2]

# Load source and template
source_wb = load_workbook(source_file, data_only=True)
source_ws = source_wb.active
b2_value = source_ws["B2"].value  # Static value

template_wb = load_workbook(template_file)
name_template = template_wb["NAME"]

# Track created operator sheets and row counters
operator_sheets = {}
row_cursor = 14  # Start at row 14 (assuming first data is at row 13)

while True:
    """
    name = source_ws[f"B{row_cursor}"].value
    if not name:
        break  # End of data
    """
    name = source_ws[f"B{row_cursor}"].value
    if not name:
        break  # End of data

    if "—" in str(name):  # skip if name contains em dash
        row_cursor += 1
        continue

    # Create or get sheet
    if name not in operator_sheets:
        new_ws = template_wb.copy_worksheet(name_template)
        new_ws.title = str(name)
        operator_sheets[name] = new_ws
        sheet_row = 13  # start from row 13
    else:
        new_ws = operator_sheets[name]
        sheet_row += 1  # move down for each new entry

    # Mapping values
    new_ws[f"C5"] = parse_number_if_numeric(source_ws[f"B{row_cursor}"].value )    # B13 ? C5
    new_ws[f"C7"] = parse_number_if_numeric(source_ws[f"A{row_cursor}"].value )    # A13 ? C7
    new_ws[f"C8"] = parse_number_if_numeric(b2_value)                              # B2  ? C8
    new_ws[f"B{sheet_row}"] = parse_number_if_numeric(source_ws[f"C{row_cursor}"].value ) # C13 ? B13
    new_ws[f"C{sheet_row}"] = parse_number_if_numeric(source_ws[f"D{row_cursor}"].value ) # D13 ? C13
    new_ws[f"D{sheet_row}"] = parse_number_if_numeric(source_ws[f"E{row_cursor}"].value)  # E13 ? D13
    new_ws[f"E{sheet_row}"] = parse_number_if_numeric(source_ws[f"H{row_cursor}"].value)  # H13 ? E13

    # I13:M13 ? F13:J13
    for i, col in enumerate(range(9, 14)):  # Columns I to M
        val = parse_number_if_numeric(source_ws[f"{get_column_letter(col)}{row_cursor}"].value)
        
        target_col = get_column_letter(6 + i)  # F=6 ? J=10
        new_ws[f"{target_col}{sheet_row}"] = val
    
    
    # Move to next row in source
    row_cursor += 1

# Remove the original "NAME" template sheet
if "NAME" in template_wb.sheetnames and len(template_wb.sheetnames) > 1:
    template_wb.remove(name_template)

# Save result
template_wb.save(output_file)
print(f"? Saved to {output_file}")